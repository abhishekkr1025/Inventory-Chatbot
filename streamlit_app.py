import streamlit as st
import openpyxl
import pandas as pd
import random
import time

st.title("Simple Chat with Inventory Management")

# Initialize chat history
if "messages" not in st.session_state:
    st.session_state.messages = []

# Display chat messages from history on app rerun
for message in st.session_state.messages:
    with st.chat_message(message["role"]):
        st.markdown(message["content"])

# Function to create or load an Excel workbook
def create_or_load_workbook(filename):
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Inventory"
        sheet.append(["Item Name", "Quantity"])
    return workbook

# Function to add or update item in the workbook
def add_or_update_item(workbook, item_name, quantity):
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=False):  # Start from the second row to skip header
        if row[0].value.lower() == item_name.lower():
            row[1].value += quantity
            workbook.save("inventory.xlsx")
            return "updated"
    sheet.append([item_name, quantity])
    workbook.save("inventory.xlsx")
    return "added"

# Function to search for an item in the workbook
def search_item_quantity(workbook, item_name):
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        if row[0].lower() == item_name.lower():
            return row[1]
    return None

# Function to generate assistant response
def response_generator():
    responses = [
        "Please provide the item name and quantity in the format 'Item: Quantity'.",
        "Got it. Now tell me the item name and quantity (e.g., 'Apples: 5').",
        "I can help you with that! What is the item name and quantity?"
    ]
    response = random.choice(responses)
    for word in response.split():
        yield word + " "
        time.sleep(0.05)

# Function to read the inventory data from the workbook
def read_inventory_data(workbook):
    sheet = workbook.active
    data = sheet.iter_rows(values_only=True)
    columns = next(data)  # Get the column names
    df = pd.DataFrame(data, columns=columns)
    return df

# Accept user input
if prompt := st.chat_input("What is up?"):
    # Display user message in chat message container
    with st.chat_message("user"):
        st.markdown(prompt)
    # Add user message to chat history
    st.session_state.messages.append({"role": "user", "content": prompt})

    # Check if user input follows the "Item: Quantity" format or is a query
    if ":" in prompt:
        try:
            item_name, quantity = prompt.split(":")
            item_name = item_name.strip()
            quantity = int(quantity.strip())

            # Load or create workbook and add or update item
            workbook = create_or_load_workbook("inventory.xlsx")
            action = add_or_update_item(workbook, item_name, quantity)

            # Generate a confirmation response
            if action == "updated":
                response = f"Updated {item_name} with additional {quantity} units."
            else:
                response = f"Added {quantity} of {item_name} to the inventory."
        except ValueError:
            response = "Invalid format. Please use 'Item: Quantity'."
    else:
        # Check if user is querying the quantity of an item
        workbook = create_or_load_workbook("inventory.xlsx")
        quantity = search_item_quantity(workbook, prompt.strip())
        if quantity is not None:
            response = f"The quantity of {prompt.strip()} in stock is: {quantity}"
        else:
            # Streamed response emulator
            response = "".join(response_generator())

    # Display assistant response in chat message container
    with st.chat_message("assistant"):
        st.markdown(response)
    # Add assistant response to chat history
    st.session_state.messages.append({"role": "assistant", "content": response})

# Load the workbook and display the inventory
workbook = create_or_load_workbook("inventory.xlsx")
inventory_df = read_inventory_data(workbook)
st.dataframe(inventory_df)
